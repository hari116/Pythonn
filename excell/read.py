import openpyxl
import xlsxwriter 

A=''

wb = openpyxl.load_workbook('test.xlsx', data_only=True)
wbF = openpyxl.load_workbook('test.xlsx')
# wb.get_sheet_names()
sheet = wb.get_sheet_by_name('Sheet1')
sheetF = wbF.get_sheet_by_name('Sheet1')
# print(sheet['A6'].value)
a = sheet['A1']
b = sheet['B1']
c = sheet['C1']

aF = sheetF['A1']
bF = sheetF['B1']
cF = sheetF['C1']
# max row col
max_row = sheet.max_row
max_col = sheet.max_column
# print(max_col)#column letter
print(a.value, b.value, c.value)
# sheet['B1']=1000
# wb.save('test.xlsx')
print(aF.value, bF.value, cF.value)
